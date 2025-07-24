"""
@Author: li
@Email: lijianqiao2906@live.com
@FileName: base.py
@DateTime: 2025/07/23
@Docs: 导入导出基础类和通用工具
"""

import tempfile
from abc import ABC, abstractmethod
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from pydantic import BaseModel

from app.core.exceptions import BusinessException
from app.utils.deps import OperationContext
from app.utils.logger import logger


class FieldMapping(BaseModel):
    """字段映射配置"""

    field_name: str  # 数据库字段名
    display_name: str  # 显示名称（中文）
    english_name: str  # 英文字段名
    is_required: bool = False  # 是否必填
    field_type: str = "str"  # 字段类型
    max_length: int | None = None  # 最大长度
    min_length: int | None = None  # 最小长度
    default_value: Any = None  # 默认值
    description: str = ""  # 字段说明
    example_value: str = ""  # 示例值
    choices: list[str] | None = None  # 选择项
    foreign_key_mapping: dict[str, str] | None = None  # 外键映射 {显示字段: 外键字段}


class ImportExportConfig(BaseModel):
    """导入导出配置"""

    model_name: str  # 模型名称
    display_name: str  # 显示名称
    sheet_name: str  # 工作表名称
    main_fields: list[FieldMapping]  # 主要字段
    foreign_key_fields: list[FieldMapping] = []  # 外键字段
    unique_fields: list[str] = []  # 唯一性校验字段
    update_fields: list[str] = []  # 允许更新的字段


class BaseImportExportService(ABC):
    """
    通用导入导出服务基类
    """

    def __init__(self, config: ImportExportConfig):
        self.config = config
        self._field_mapping = {}
        self._reverse_mapping = {}
        self._init_mapping()

    def _init_mapping(self):
        """初始化字段映射"""
        all_fields = self.config.main_fields + self.config.foreign_key_fields

        for field in all_fields:
            # 中文名 -> 英文字段名
            self._field_mapping[field.display_name] = field.field_name
            self._field_mapping[field.english_name] = field.field_name

            # 英文字段名 -> 中文名
            self._reverse_mapping[field.field_name] = field.display_name

    def _build_headers(self, include_required_mark: bool = True) -> list[str]:
        """构建表头"""
        headers = []
        all_fields = self.config.main_fields + self.config.foreign_key_fields

        for field in all_fields:
            if include_required_mark and field.is_required:
                headers.append(f"{field.display_name}*")
            else:
                headers.append(field.display_name)

        return headers

    def _set_header_style(self, sheet, color: str = "366092"):
        """设置标题行样式"""
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    def _adjust_column_width(self, sheet):
        """自动调整列宽"""
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width

    def _create_temp_file(self, suffix: str, prefix: str) -> str:
        """创建临时文件并返回路径"""
        with tempfile.NamedTemporaryFile(suffix=suffix, prefix=prefix, delete=False) as temp_file:
            return temp_file.name

    # ================= 导出功能 =================

    async def export_template(self, file_format: str = "xlsx") -> str:
        """导出导入模板"""
        try:
            if file_format.lower() == "xlsx":
                return await self._create_excel_template()
            elif file_format.lower() == "csv":
                return await self._create_csv_template()
            else:
                raise BusinessException(f"不支持的文件格式: {file_format}")
        except Exception as e:
            logger.error(f"导出模板失败: {e}")
            raise BusinessException(f"导出模板失败: {e}") from e

    async def _create_excel_template(self) -> str:
        """创建Excel导入模板"""
        wb = Workbook()

        # 删除默认工作表
        wb.remove(wb.active)  # type: ignore

        # 创建示例数据工作表
        example_sheet = wb.create_sheet("示例数据")
        await self._create_example_sheet(example_sheet)

        # 创建字段说明工作表
        field_sheet = wb.create_sheet("字段说明")
        await self._create_field_description_sheet(field_sheet)

        # 创建导入模板工作表
        template_sheet = wb.create_sheet("导入模板")
        await self._create_import_template_sheet(template_sheet)

        # 保存到临时文件
        filepath = self._create_temp_file(".xlsx", f"{self.config.model_name}_import_template_")

        # 保存文件并确保工作簿正确关闭
        wb.save(filepath)
        wb.close()

        return filepath

    async def _create_example_sheet(self, sheet):
        """创建示例数据工作表"""
        # 设置标题
        headers = self._build_headers(include_required_mark=True)
        sheet.append(headers)

        # 设置标题样式
        self._set_header_style(sheet, "366092")

        # 添加示例数据
        example_data = await self._get_example_data()
        for row_data in example_data:
            sheet.append(row_data)

        # 自动调整列宽
        self._adjust_column_width(sheet)

    async def _create_field_description_sheet(self, sheet):
        """创建字段说明工作表"""
        headers = ["字段名称", "英文字段名", "是否必填", "字段类型", "最大长度", "描述", "示例值", "选择项"]
        sheet.append(headers)

        # 设置标题样式
        self._set_header_style(sheet, "70AD47")

        all_fields = self.config.main_fields + self.config.foreign_key_fields

        for field in all_fields:
            row = [
                field.display_name,
                field.english_name,
                "是" if field.is_required else "否",
                field.field_type,
                field.max_length or "",
                field.description,
                field.example_value,
                ", ".join(field.choices) if field.choices else "",
            ]
            sheet.append(row)

        # 自动调整列宽
        self._adjust_column_width(sheet)

    async def _create_import_template_sheet(self, sheet):
        """创建导入模板工作表"""
        headers = self._build_headers(include_required_mark=True)
        sheet.append(headers)

        # 设置标题样式
        self._set_header_style(sheet, "C5504B")

        # 自动调整列宽
        all_fields = self.config.main_fields + self.config.foreign_key_fields
        for i, field in enumerate(all_fields, 1):
            column_letter = sheet.cell(row=1, column=i).column_letter
            width = max(len(field.display_name) + 2, 15)
            sheet.column_dimensions[column_letter].width = width

    async def _create_csv_template(self) -> str:
        """创建CSV导入模板"""
        headers = self._build_headers(include_required_mark=True)
        df = pd.DataFrame(data=None, columns=pd.Index(headers))

        # 添加示例数据
        example_data = await self._get_example_data()
        for row_data in example_data:
            df.loc[len(df)] = row_data

        # 保存到临时文件
        filepath = self._create_temp_file(".csv", f"{self.config.model_name}_import_template_")

        # 保存CSV文件
        df.to_csv(filepath, index=False, encoding="utf-8-sig")

        return filepath

    async def export_data(
        self, data: list[Any], file_format: str = "xlsx", operation_context: OperationContext | None = None
    ) -> str:
        """导出数据"""
        try:
            if file_format.lower() == "xlsx":
                return await self._export_excel_data(data)
            elif file_format.lower() == "csv":
                return await self._export_csv_data(data)
            else:
                raise BusinessException(f"不支持的文件格式: {file_format}")
        except Exception as e:
            logger.error(f"导出数据失败: {e}")
            raise BusinessException(f"导出数据失败: {e}") from e

    async def _export_excel_data(self, data: list[Any]) -> str:
        """导出Excel数据"""
        headers = self._build_headers(include_required_mark=False)

        # 转换数据
        rows = []
        for item in data:
            row = await self._convert_model_to_export_row(item)
            rows.append(row)

        df = pd.DataFrame(rows, columns=pd.Index(headers))

        # 创建临时文件
        filepath = self._create_temp_file(".xlsx", f"{self.config.model_name}_export_")

        # 保存Excel文件
        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=self.config.sheet_name, index=False)

            # 设置样式
            workbook = writer.book  # noqa: F841
            worksheet = writer.sheets[self.config.sheet_name]

            # 设置标题样式
            self._set_header_style(worksheet, "366092")

            # 自动调整列宽
            self._adjust_column_width(worksheet)

        return filepath

    async def _export_csv_data(self, data: list[Any]) -> str:
        """导出CSV数据"""
        headers = self._build_headers(include_required_mark=False)

        # 转换数据
        rows = []
        for item in data:
            row = await self._convert_model_to_export_row(item)
            rows.append(row)

        df = pd.DataFrame(rows, columns=pd.Index(headers))

        # 创建临时文件
        filepath = self._create_temp_file(".csv", f"{self.config.model_name}_export_")

        # 保存CSV文件
        df.to_csv(filepath, index=False, encoding="utf-8-sig")

        return filepath

    # ================= 导入功能 =================

    async def import_data(
        self, file_path: str, operation_context: OperationContext, update_existing: bool = False
    ) -> dict[str, Any]:
        """导入数据"""
        try:
            # 读取文件
            if file_path.endswith(".xlsx"):
                df = await self._read_excel_file(file_path)
            elif file_path.endswith(".csv"):
                df = await self._read_csv_file(file_path)
            else:
                raise BusinessException("只支持 .xlsx 和 .csv 格式文件")

            # 验证数据
            validation_result = await self._validate_import_data(df)
            if validation_result["errors"]:
                return {
                    "success": False,
                    "total_rows": len(df),
                    "success_count": 0,
                    "error_count": len(df),
                    "errors": validation_result["errors"],
                    "imported_ids": [],
                }

            # 转换数据
            converted_data = await self._convert_import_data(validation_result["valid_data"])

            # 导入数据
            import_result = await self._import_validated_data(converted_data, operation_context, update_existing)

            return {
                "success": True,
                "total_rows": len(df),
                "success_count": import_result["success_count"],
                "error_count": import_result["error_count"],
                "errors": import_result["errors"],
                "imported_ids": import_result["imported_ids"],
            }

        except Exception as e:
            logger.error(f"导入数据失败: {e}")
            raise BusinessException(f"导入数据失败: {e}") from e

    async def _read_excel_file(self, file_path: str) -> pd.DataFrame:
        """读取Excel文件"""
        try:
            # 使用上下文管理器确保文件正确关闭
            with pd.ExcelFile(file_path) as excel_file:
                logger.info(f"Excel文件工作表列表: {excel_file.sheet_names}")

                # 优先读取"导入模板"工作表
                if "导入模板" in excel_file.sheet_names:
                    logger.info("读取'导入模板'工作表")
                    df = excel_file.parse("导入模板")
                elif self.config.sheet_name in excel_file.sheet_names:
                    logger.info(f"读取'{self.config.sheet_name}'工作表")
                    df = excel_file.parse(self.config.sheet_name)
                else:
                    # 读取第一个工作表
                    logger.info(f"读取第一个工作表: {excel_file.sheet_names[0]}")
                    df = excel_file.parse(0)

            # 文件已在with语句中关闭，现在可以安全返回数据
            # 确保返回的是DataFrame类型
            if isinstance(df, pd.DataFrame):
                return df
            elif isinstance(df, dict):
                # 如果是字典类型，返回第一个DataFrame
                first_df = next(iter(df.values()))
                if isinstance(first_df, pd.DataFrame):
                    return first_df
                else:
                    raise BusinessException("无法获取有效的DataFrame数据")
            else:
                raise BusinessException("无法获取有效的DataFrame数据")
        except Exception as e:
            logger.error(f"读取Excel文件失败，文件路径: {file_path}")
            raise BusinessException(f"读取Excel文件失败: {e}") from e

    async def _read_csv_file(self, file_path: str) -> pd.DataFrame:
        """读取CSV文件"""
        try:
            df = pd.read_csv(file_path, encoding="utf-8-sig")
            return df
        except UnicodeDecodeError:
            try:
                df = pd.read_csv(file_path, encoding="gbk")
                return df
            except Exception as e:
                raise BusinessException(f"读取CSV文件失败，请检查文件编码: {e}") from e
        except Exception as e:
            raise BusinessException(f"读取CSV文件失败: {e}") from e

    async def _validate_import_data(self, df: pd.DataFrame) -> dict[str, Any]:
        """验证导入数据"""
        errors = []
        valid_data = []

        # 检查是否为空
        if df.empty:
            logger.warning("数据框为空")
            errors.append("文件内容为空")
            return {"errors": errors, "valid_data": valid_data}

        # 标准化列名（去除*号和空格）
        df.columns = [col.rstrip("*").strip() for col in df.columns]

        # 检查必填字段是否存在
        required_fields = [
            field.display_name
            for field in self.config.main_fields + self.config.foreign_key_fields
            if field.is_required
        ]

        missing_fields = [field for field in required_fields if field not in df.columns]
        if missing_fields:
            logger.warning(f"缺少必填字段: {missing_fields}")
            errors.append(f"缺少必填字段: {', '.join(missing_fields)}")
            return {"errors": errors, "valid_data": valid_data}

        # 逐行验证数据
        for row_idx, (_, row) in enumerate(df.iterrows()):
            row_errors = []
            row_number = row_idx + 2  # Excel行号（考虑标题行），从第2行开始

            # 检查必填字段
            for field in self.config.main_fields + self.config.foreign_key_fields:
                if field.is_required:
                    value = row.get(field.display_name)
                    if pd.isna(value) or str(value).strip() == "":  # pyright: ignore[reportGeneralTypeIssues]
                        row_errors.append(f"第{row_number}行：{field.display_name} 为必填字段")

            # 字段格式验证
            field_validation_errors = await self._validate_row_fields(row.to_dict(), row_number)
            row_errors.extend(field_validation_errors)

            if row_errors:
                errors.extend(row_errors)
            else:
                valid_data.append(row.to_dict())

        return {"errors": errors, "valid_data": valid_data}

    def _safe_str_convert(self, value: Any, default: str = "", allow_empty: bool = False) -> str | None:
        """安全的字符串转换，处理NaN值"""
        if pd.isna(value) or value is None:
            if allow_empty:
                return None
            return default

        str_value = str(value).strip()
        if not str_value and allow_empty:
            return None

        return str_value if str_value else default

    def _safe_int_convert(self, value: Any, default: int = 0) -> int:
        """安全的整数转换，处理NaN值"""
        if pd.isna(value) or value is None or value == "":
            return default

        try:
            return int(value)
        except (ValueError, TypeError):
            return default

    # ================= 抽象方法 =================

    @abstractmethod
    async def _get_example_data(self) -> list[list[str]]:
        """获取示例数据"""
        pass

    @abstractmethod
    async def _convert_model_to_export_row(self, model: Any) -> list[str]:
        """将模型转换为导出行数据"""
        pass

    @abstractmethod
    async def _validate_row_fields(self, row: dict[str, Any], row_number: int) -> list[str]:
        """验证行字段格式"""
        pass

    @abstractmethod
    async def _convert_import_data(self, data: list[dict[str, Any]]) -> list[dict[str, Any]]:
        """转换导入数据格式"""
        pass

    @abstractmethod
    async def _import_validated_data(
        self, data: list[dict[str, Any]], operation_context: OperationContext, update_existing: bool
    ) -> dict[str, Any]:
        """导入验证通过的数据"""
        pass
